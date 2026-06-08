"""
excel_gen.py — Exporta una cotización a archivo .xlsx para clientes que
piden la tabla editable. Hojas: Resumen + Letras + Desglose.
"""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from calculator import QuoteResult

_HEAD_FILL  = PatternFill("solid", fgColor="1F3A5F")
_HEAD_FONT  = Font(bold=True, color="FFFFFF", size=11)
_TOTAL_FONT = Font(bold=True, size=11)
_CENTER     = Alignment(horizontal="center", vertical="center")
_LEFT       = Alignment(horizontal="left", vertical="center")
_RIGHT      = Alignment(horizontal="right", vertical="center")


def _header_row(ws, row: int, headers: list[str]) -> None:
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.fill = _HEAD_FILL
        cell.font = _HEAD_FONT
        cell.alignment = _CENTER


def _autosize(ws, n_cols: int) -> None:
    for col in range(1, n_cols + 1):
        letter = get_column_letter(col)
        max_len = 0
        for row in ws[letter]:
            v = row.value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[letter].width = min(max_len + 2, 50)


def generar_xlsx(result: QuoteResult, meta: dict) -> bytes:
    """Genera el .xlsx en memoria y devuelve los bytes."""
    wb = Workbook()

    # ─── Hoja 1: Resumen ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumen"

    ws["A1"] = "COTIZACIÓN SGI"
    ws["A1"].font = Font(bold=True, size=16, color="1F3A5F")
    ws.merge_cells("A1:D1")
    ws["A1"].alignment = _CENTER

    info = [
        ("Folio",   meta.get("folio", "—")),
        ("Cliente", meta.get("cliente", "—")),
        ("Fecha",   meta.get("fecha", "—")),
        ("Tipo",    result.tipo),
        ("Notas",   meta.get("notas", "")),
    ]
    for i, (k, v) in enumerate(info, start=3):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)

    base_row = 3 + len(info) + 1
    ws.cell(row=base_row, column=1, value="MEDIDAS").font = Font(bold=True, color="1F3A5F")
    medidas = [
        ("Paths detectados",     result.paths_count),
        ("Altura de letra (cm)", round(result.altura_letra_cm, 1)),
        ("Área cara (cm²)",      round(result.area_cara_cm2, 2)),
        ("Perímetro total (cm)", round(result.perimetro_total_cm, 2)),
        ("Cercha altura (cm)",   round(result.cercha_altura_cm, 1)),
    ]
    for i, (k, v) in enumerate(medidas, start=base_row + 1):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v).alignment = _RIGHT

    costo_row = base_row + len(medidas) + 2
    ws.cell(row=costo_row, column=1, value="COSTOS").font = Font(bold=True, color="1F3A5F")
    costos = [
        ("Subtotal",             round(result.subtotal, 2)),
        ("IVA (16%)",            round(result.iva, 2)),
        ("Total",                round(result.total, 2)),
        ("Precio sin ajuste",    round(result.precio_sin_ajuste, 2)),
        ("Ajuste (%)",           round(result.ajuste_pct, 2)),
        ("Precio venta sugerido", round(result.precio_venta_sugerido, 2)),
        ("Precio piso por costo", round(result.precio_venta_costo, 2)),
        ("Instalación",          round(result.inst_total, 2)),
        ("PRECIO FINAL",         round(result.precio_final, 2)),
    ]
    for i, (k, v) in enumerate(costos, start=costo_row + 1):
        c1 = ws.cell(row=i, column=1, value=k)
        c2 = ws.cell(row=i, column=2, value=v)
        c2.alignment = _RIGHT
        c2.number_format = '"$" #,##0.00'
        if k == "PRECIO FINAL":
            c1.font = _TOTAL_FONT
            c2.font = _TOTAL_FONT
            c1.fill = PatternFill("solid", fgColor="FFF4D6")
            c2.fill = PatternFill("solid", fgColor="FFF4D6")

    _autosize(ws, 2)

    # ─── Hoja 2: Letras (desglose por pieza) ─────────────────────────────────
    if result.desglose_letras:
        ws2 = wb.create_sheet("Letras")
        headers = ["ID", "Alto (cm)", "Ancho (cm)", "Área bbox (cm²)",
                   "Perímetro (cm)", "Cercha área (cm²)",
                   "Costo cara", "Costo cercha", "Costo material",
                   "Precio letra"]
        _header_row(ws2, 1, headers)
        for i, d in enumerate(result.desglose_letras, start=2):
            row_vals = [
                d.get("id", ""),
                d.get("alto_cm", 0),
                d.get("ancho_cm", 0),
                d.get("area_bbox_cm2", 0),
                d.get("perimetro_cm", 0),
                d.get("cercha_area_cm2", 0),
                d.get("costo_cara", 0),
                d.get("costo_cercha", 0),
                d.get("costo_mat", 0),
                d.get("precio_letra", 0),
            ]
            for col, v in enumerate(row_vals, start=1):
                cell = ws2.cell(row=i, column=col, value=v)
                if col >= 7:  # columnas de dinero
                    cell.number_format = '"$" #,##0.00'
                    cell.alignment = _RIGHT
                elif col >= 2:
                    cell.alignment = _RIGHT
        # Fila de totales
        last = len(result.desglose_letras) + 2
        ws2.cell(row=last, column=1, value="TOTAL").font = _TOTAL_FONT
        for col in (7, 8, 9, 10):
            ws2.cell(row=last, column=col,
                     value=f"=SUM({get_column_letter(col)}2:{get_column_letter(col)}{last-1})")
            ws2.cell(row=last, column=col).number_format = '"$" #,##0.00'
            ws2.cell(row=last, column=col).font = _TOTAL_FONT
        _autosize(ws2, len(headers))

    # ─── Hoja 3: Desglose de costos ──────────────────────────────────────────
    if result.desglose:
        ws3 = wb.create_sheet("Desglose")
        _header_row(ws3, 1, ["Concepto", "Costo"])
        for i, d in enumerate(result.desglose, start=2):
            ws3.cell(row=i, column=1, value=d.get("concepto", ""))
            c = ws3.cell(row=i, column=2, value=d.get("costo", 0))
            c.number_format = '"$" #,##0.00'
            c.alignment = _RIGHT
        _autosize(ws3, 2)

    # Guardar a buffer
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
